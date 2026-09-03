"""Collect raw point-in-time series required by MRS-v0.1."""

from __future__ import annotations

import csv
from datetime import date, datetime
from html import unescape
from io import BytesIO, StringIO
import re
import ssl
import time
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

import certifi
from openpyxl import load_workbook
from pypdf import PdfReader


JST = ZoneInfo("Asia/Tokyo")
TLS_CONTEXT = ssl.create_default_context(cafile=certifi.where())


class MarketRegimeSourceError(RuntimeError):
    pass


def _download(url: str, *, timeout: int, attempts: int = 3) -> tuple[bytes, str]:
    request = Request(
        url,
        headers={
            "User-Agent": "stock.jp research operation/1.0 (+https://github.com/kkyosuke/stock.jp)",
            "Accept": "application/json,text/csv,text/html,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    for attempt in range(attempts):
        try:
            with urlopen(request, timeout=timeout, context=TLS_CONTEXT) as response:
                return response.read(), response.geturl()
        except HTTPError as error:
            if error.code not in {429, 500, 502, 503, 504} or attempt + 1 == attempts:
                raise MarketRegimeSourceError(
                    f"HTTP {error.code} from {url}"
                ) from error
        except URLError as error:
            if attempt + 1 == attempts:
                raise MarketRegimeSourceError(
                    f"network error from {url}: {error.reason}"
                ) from error
        time.sleep(0.5 * (2**attempt))
    raise AssertionError("unreachable")


def _jpx_pdf_links(page_urls: list[str], *, timeout: int) -> list[str]:
    links: set[str] = set()
    for page_url in page_urls:
        page_bytes, final_page_url = _download(page_url, timeout=timeout)
        try:
            page = page_bytes.decode("utf-8")
        except UnicodeDecodeError:
            page = page_bytes.decode("shift_jis")
        for match in re.finditer(
            r'href=["\']([^"\']*/03_sisu\d{4}\.pdf)["\']', page, flags=re.IGNORECASE
        ):
            links.add(urljoin(final_page_url, unescape(match.group(1))))
    return sorted(links)


def _jpx_month_points(
    payload: bytes, url: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    match = re.search(r"03_sisu(\d{2})(\d{2})\.pdf", url)
    if not match:
        raise MarketRegimeSourceError(f"cannot identify year/month from {url}")
    year, month = 2000 + int(match.group(1)), int(match.group(2))
    try:
        text = PdfReader(BytesIO(payload)).pages[0].extract_text()
    except Exception as error:
        raise MarketRegimeSourceError(
            f"cannot parse JPX monthly index PDF {url}"
        ) from error
    lines = text.splitlines()
    start = next(
        (
            index
            for index, line in enumerate(lines)
            if re.match(rf"^{month}\.\d{{1,2}}(?:\s+|$)", line.strip())
        ),
        None,
    )
    if start is None:
        raise MarketRegimeSourceError(f"daily index table was not found in {url}")
    topix: list[dict[str, Any]] = []
    growth: list[dict[str, Any]] = []
    for line in lines[start:]:
        stripped = line.strip()
        if stripped.startswith("平均"):
            break
        first = re.match(rf"^(?:{month}\.)?(\d{{1,2}})(?:\s+|$)", stripped)
        if not first:
            continue
        day = int(first.group(1))
        try:
            session = date(year, month, day)
        except ValueError:
            continue
        numbers = [
            float(value.replace(",", ""))
            for value in re.findall(r"\d[\d,]*\.\d+", stripped[first.end() :])
        ]
        if len(numbers) < 2:
            continue
        # JPX table 3-1 places TOPIX first and Growth Market 250 last.
        topix.append({"date": session.isoformat(), "value": numbers[0]})
        growth.append({"date": session.isoformat(), "value": numbers[-1]})
    if not topix:
        raise MarketRegimeSourceError(f"no daily index values parsed from {url}")
    return topix, growth


def _jpx_index_points(
    page_urls: list[str], *, as_of: date, timeout: int
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    links = _jpx_pdf_links(page_urls, timeout=timeout)
    applicable = [
        url
        for url in links
        if (match := re.search(r"03_sisu(\d{2})(\d{2})\.pdf", url))
        and date(2000 + int(match.group(1)), int(match.group(2)), 1)
        <= as_of.replace(day=1)
    ]
    applicable.sort(
        key=lambda url: tuple(
            map(int, re.search(r"03_sisu(\d{2})(\d{2})\.pdf", url).groups())
        )
    )
    selected = applicable[-14:]
    if not selected:
        raise MarketRegimeSourceError(
            "JPX monthly statistics contain no applicable index PDFs"
        )
    topix_by_date: dict[str, dict[str, Any]] = {}
    growth_by_date: dict[str, dict[str, Any]] = {}
    used_urls: list[str] = []
    for url in selected:
        payload, final_url = _download(url, timeout=timeout)
        topix, growth = _jpx_month_points(payload, final_url)
        topix_by_date.update(
            {
                point["date"]: point
                for point in topix
                if date.fromisoformat(point["date"]) <= as_of
            }
        )
        growth_by_date.update(
            {
                point["date"]: point
                for point in growth
                if date.fromisoformat(point["date"]) <= as_of
            }
        )
        used_urls.append(final_url)
    topix_points = [topix_by_date[key] for key in sorted(topix_by_date)]
    growth_points = [growth_by_date[key] for key in sorted(growth_by_date)]
    if len(topix_points) < 200 or len(growth_points) < 200:
        raise MarketRegimeSourceError(
            f"JPX monthly PDFs yielded TOPIX={len(topix_points)}, Growth250={len(growth_points)} observations"
        )
    return topix_points, growth_points, used_urls


def _nikkei_vi_points(
    url: str, *, as_of: date, timeout: int
) -> tuple[list[dict[str, Any]], str]:
    payload, final_url = _download(url, timeout=timeout)
    try:
        try:
            decoded = payload.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = payload.decode("cp932")
        rows = csv.DictReader(StringIO(decoded))
        points = [
            {
                "date": datetime.strptime(row["Date of Data"], "%Y/%m/%d")
                .date()
                .isoformat(),
                "value": float(row["Close"]),
            }
            for row in rows
            if row.get("Close")
            and datetime.strptime(row["Date of Data"], "%Y/%m/%d").date() <= as_of
        ]
    except (KeyError, UnicodeDecodeError, ValueError) as error:
        raise MarketRegimeSourceError("invalid Nikkei VI daily CSV") from error
    if not points:
        raise MarketRegimeSourceError("Nikkei VI daily CSV contained no usable rows")
    return points, final_url


def _leading_ci(
    page_url: str, *, as_of: date, timeout: int
) -> tuple[list[dict[str, Any]], str, str]:
    page_bytes, _ = _download(page_url, timeout=timeout)
    try:
        page = page_bytes.decode("utf-8")
    except UnicodeDecodeError:
        page = page_bytes.decode("shift_jis")
    matches = list(
        re.finditer(r'href=["\']([^"\']*\d{4}ci\.xlsx)["\']', page, flags=re.IGNORECASE)
    )
    if not matches:
        raise MarketRegimeSourceError(
            "Cabinet Office page has no long-series CI workbook"
        )
    selected = matches[0]
    workbook_url = urljoin(page_url, unescape(selected.group(1)))
    context = unescape(
        re.sub(
            r"<[^>]+>", " ", page[max(0, selected.start() - 3000) : selected.start()]
        )
    )
    published_dates = re.findall(r"(20\d{2})\)?年\s*(\d{1,2})月\s*(\d{1,2})日", context)
    if not published_dates:
        raise MarketRegimeSourceError(
            "Cabinet Office CI publication date was not found"
        )
    published = date(*map(int, published_dates[-1]))
    if published > as_of:
        raise MarketRegimeSourceError(
            f"latest Cabinet Office CI workbook was published {published}, after MRS as_of {as_of}"
        )
    workbook_bytes, final_url = _download(workbook_url, timeout=timeout)
    try:
        workbook = load_workbook(
            BytesIO(workbook_bytes), read_only=True, data_only=True
        )
        sheet = workbook["指数 Indexes"]
        points = []
        for row in sheet.iter_rows(min_row=7, values_only=True):
            year, month, leading = row[1], row[2], row[3]
            if year is None or month is None or leading is None:
                continue
            period = date(int(year), int(month), 1)
            if period <= as_of:
                points.append(
                    {"period": period.strftime("%Y-%m"), "value": float(leading)}
                )
    except (KeyError, TypeError, ValueError) as error:
        raise MarketRegimeSourceError("invalid Cabinet Office CI workbook") from error
    if not points:
        raise MarketRegimeSourceError(
            "Cabinet Office CI workbook contained no usable rows"
        )
    return (
        points,
        final_url,
        datetime.combine(published, datetime.min.time(), tzinfo=JST).isoformat(
            timespec="seconds"
        ),
    )


def collect_market_regime_series(
    *, as_of: date, config: dict[str, Any]
) -> dict[str, Any]:
    """Download raw inputs; any missing provider fails the whole formal MRS."""

    timeout = int(config.get("request_timeout_seconds", 30))
    market = config.get("market_regime", {})
    if not isinstance(market, dict) or market.get("enabled") is not True:
        raise MarketRegimeSourceError("market_regime collection is disabled")
    pages = market.get("jpx_monthly_statistics_pages")
    if not isinstance(pages, list) or not pages:
        raise MarketRegimeSourceError(
            "market_regime.jpx_monthly_statistics_pages is empty"
        )
    topix, growth, jpx_urls = _jpx_index_points(
        [str(value) for value in pages], as_of=as_of, timeout=timeout
    )
    vi, vi_url = _nikkei_vi_points(
        str(market["nikkei_vi_csv_url"]), as_of=as_of, timeout=timeout
    )
    ci, ci_url, ci_available = _leading_ci(
        str(market["leading_ci_page_url"]), as_of=as_of, timeout=timeout
    )
    return {
        "schema_version": "1.0",
        "as_of": as_of.isoformat(),
        "minimum_vi_observations": int(market.get("minimum_vi_observations", 500)),
        "series": {
            "topix": {
                "source_url": "|".join(jpx_urls),
                "primary_source": True,
                "points": topix,
            },
            "growth250": {
                "source_url": "|".join(jpx_urls),
                "primary_source": True,
                "points": growth,
            },
            "nikkei_vi": {"source_url": vi_url, "primary_source": True, "points": vi},
            "leading_ci": {
                "source_url": ci_url,
                "primary_source": True,
                "available_at_jst": ci_available,
                "points": ci,
            },
        },
    }
