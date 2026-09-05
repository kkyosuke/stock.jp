"""Collect non-EDINET official reference data for a nightly operation.

Raw and normalized records are intentionally written below ``operations/private``.
J-Quants data must not be copied into the public repository.
"""

from __future__ import annotations

import csv
from datetime import date, datetime, time, timedelta
import gzip
import hashlib
import io
import json
import math
import os
from pathlib import Path
import ssl
import statistics
import tempfile
import time as time_module
from collections.abc import Mapping
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

import certifi
from openpyxl import load_workbook


JST = ZoneInfo("Asia/Tokyo")
TLS_CONTEXT = ssl.create_default_context(cafile=certifi.where())
USER_AGENT = "Mozilla/5.0 (compatible; stock.jp nightly research/0.1)"

MASTER_FIELDS = (
    "effective_date",
    "code",
    "company_name",
    "company_name_en",
    "market_code",
    "market_name",
    "sector_17_code",
    "sector_17_name",
    "sector_33_code",
    "sector_33_name",
    "scale_category",
    "margin_category",
    "margin_category_name",
    "product_category",
    "provider",
    "retrieved_at_jst",
)
BAR_FIELDS = (
    "date",
    "code",
    "open",
    "high",
    "low",
    "close",
    "volume",
    "trading_value",
    "adjustment_factor",
    "adjusted_close",
    "adjusted_volume",
    "provider",
)
LIQUIDITY_FIELDS = (
    "as_of_date",
    "code",
    "sessions",
    "average_trading_value_20d",
    "median_trading_value_20d",
    "latest_close",
    "latest_volume",
    "latest_trading_value",
    "provider",
    "evidence_path",
)
FORECAST_FIELDS = (
    "source_id",
    "disclosed_at_jst",
    "code",
    "disclosure_number",
    "document_type",
    "current_period_type",
    "current_fiscal_year_end",
    "forecast_sales",
    "forecast_operating_profit",
    "forecast_ordinary_profit",
    "forecast_net_profit",
    "forecast_eps",
    "next_forecast_sales",
    "next_forecast_operating_profit",
    "next_forecast_ordinary_profit",
    "next_forecast_net_profit",
    "next_forecast_eps",
    "material_change_submitted",
    "retrieved_at_jst",
    "evidence_path",
)
SHARE_FIELDS = (
    "source_id",
    "disclosed_at_jst",
    "fiscal_year_end",
    "code",
    "issued_shares_fy",
    "treasury_shares_fy",
    "average_shares",
    "retrieved_at_jst",
    "evidence_path",
    "notes",
)
EARNINGS_FIELDS = (
    "source_id",
    "published_date",
    "scheduled_date",
    "fiscal_quarter",
    "fiscal_year_end_mmdd",
    "code",
    "company_name",
    "company_name_en",
    "retrieved_at_jst",
    "evidence_path",
)


class NonEdinetSourceError(RuntimeError):
    pass


def _atomic_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w", encoding="utf-8", dir=path.parent, delete=False
    ) as temporary:
        json.dump(value, temporary, ensure_ascii=False, indent=2)
        temporary.write("\n")
        temporary_path = Path(temporary.name)
    temporary_path.chmod(0o600)
    temporary_path.replace(path)


def _atomic_csv(path: Path, fields: tuple[str, ...], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w", encoding="utf-8", newline="", dir=path.parent, delete=False
    ) as temporary:
        writer = csv.DictWriter(temporary, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
        temporary_path = Path(temporary.name)
    temporary_path.chmod(0o600)
    temporary_path.replace(path)


def _atomic_csv_gzip(
    path: Path, fields: tuple[str, ...], rows: list[dict[str, str]]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("wb", dir=path.parent, delete=False) as temporary:
        with gzip.GzipFile(fileobj=temporary, mode="wb", mtime=0) as compressed:
            with io.TextIOWrapper(compressed, encoding="utf-8", newline="") as text:
                writer = csv.DictWriter(text, fieldnames=fields, lineterminator="\n")
                writer.writeheader()
                writer.writerows(rows)
        temporary_path = Path(temporary.name)
    temporary_path.chmod(0o600)
    temporary_path.replace(path)


def _append_csv_unique(
    path: Path,
    fields: tuple[str, ...],
    rows: list[dict[str, str]],
    key: str,
) -> int:
    with path.open(encoding="utf-8", newline="") as source:
        existing = list(csv.DictReader(source))
    known = {row.get(key, "") for row in existing}
    added: list[dict[str, str]] = []
    for row in rows:
        value = row.get(key, "")
        if not value or value in known:
            continue
        known.add(value)
        added.append(row)
    if not added:
        return 0
    _atomic_csv(path, fields, existing + added)
    return len(added)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _stable_id(*values: str) -> str:
    return hashlib.sha256("|".join(values).encode("utf-8")).hexdigest()[:16]


def _normalize_code(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    code = str(value or "").strip()
    if len(code) == 5 and code.endswith("0") and code.isdigit():
        return code[:4]
    return code


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value).replace("/", "-")
    if len(text) == 8 and text.isdigit():
        return f"{text[:4]}-{text[4:6]}-{text[6:]}"
    return text


def _disclosed_at(row: Mapping[str, Any]) -> str:
    day = _date_text(row.get("DiscDate"))
    raw_time = _text(row.get("DiscTime")) or "00:00:00"
    if len(raw_time) == 4 and raw_time.isdigit():
        raw_time = f"{raw_time[:2]}:{raw_time[2:]}:00"
    elif len(raw_time) == 5:
        raw_time += ":00"
    return datetime.fromisoformat(f"{day}T{raw_time}").replace(tzinfo=JST).isoformat(
        timespec="seconds"
    )


def _request_json(
    *,
    base_url: str,
    path: str,
    params: dict[str, str],
    api_key: str,
    timeout: int,
) -> dict[str, Any]:
    public_url = f"{base_url.rstrip('/')}/{path.lstrip('/')}"
    request = Request(
        f"{public_url}?{urlencode(params)}" if params else public_url,
        headers={"x-api-key": api_key, "Accept": "application/json", "User-Agent": USER_AGENT},
    )
    for attempt in range(3):
        try:
            with urlopen(request, timeout=timeout, context=TLS_CONTEXT) as response:
                payload = json.loads(response.read().decode("utf-8"))
            if not isinstance(payload, dict):
                raise NonEdinetSourceError(f"unexpected JSON shape from {public_url}")
            return payload
        except HTTPError as error:
            retryable = error.code == 429 or error.code >= 500
            if not retryable or attempt == 2:
                raise NonEdinetSourceError(
                    f"HTTP {error.code} from {public_url}"
                ) from error
        except (URLError, TimeoutError) as error:
            if attempt == 2:
                raise NonEdinetSourceError(
                    f"network error from {public_url}: {error}"
                ) from error
        except (UnicodeDecodeError, json.JSONDecodeError) as error:
            raise NonEdinetSourceError(f"invalid JSON from {public_url}") from error
        time_module.sleep(0.5 * (2**attempt))
    raise AssertionError("retry loop did not return")


def _jquants_pages(
    *,
    base_url: str,
    path: str,
    params: dict[str, str],
    api_key: str,
    timeout: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    pages: list[dict[str, Any]] = []
    query = dict(params)
    for _ in range(100):
        payload = _request_json(
            base_url=base_url,
            path=path,
            params=query,
            api_key=api_key,
            timeout=timeout,
        )
        pages.append(payload)
        data = payload.get("data", [])
        if not isinstance(data, list):
            raise NonEdinetSourceError(f"unexpected data field from {path}")
        rows.extend(row for row in data if isinstance(row, dict))
        pagination_key = payload.get("pagination_key")
        if not pagination_key:
            return rows, pages
        query["pagination_key"] = str(pagination_key)
    raise NonEdinetSourceError(f"pagination limit exceeded for {path}")


def _fixture_pages(fixture_dir: Path, stem: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    path = fixture_dir / f"{stem}.json"
    if not path.is_file():
        raise NonEdinetSourceError(f"fixture missing: {path.name}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("data", []), list):
        raise NonEdinetSourceError(f"fixture {path.name} has invalid data")
    return [row for row in payload.get("data", []) if isinstance(row, dict)], [payload]


def _download_jpx_list(url: str, destination: Path, timeout: int) -> None:
    request = Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urlopen(request, timeout=timeout, context=TLS_CONTEXT) as response:
            body = response.read()
    except (HTTPError, URLError, TimeoutError) as error:
        raise NonEdinetSourceError(f"JPX listed master download failed: {error}") from error
    if len(body) < 10_000:
        raise NonEdinetSourceError("JPX listed master was unexpectedly small")
    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("wb", dir=destination.parent, delete=False) as temporary:
        temporary.write(body)
        temporary_path = Path(temporary.name)
    temporary_path.chmod(0o600)
    temporary_path.replace(destination)


def _parse_jpx_list(path: Path, retrieved_at: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook.worksheets[0].iter_rows(values_only=True)
        headers = [_text(value) for value in next(rows)]
        result: list[dict[str, str]] = []
        for values in rows:
            source = dict(zip(headers, values))
            market_name = _text(source.get("市場・商品区分"))
            if "内国株式" not in market_name:
                continue
            code = _normalize_code(source.get("コード"))
            if not code:
                continue
            result.append(
                {
                    "effective_date": _date_text(source.get("日付")),
                    "code": code,
                    "company_name": _text(source.get("銘柄名")),
                    "company_name_en": "",
                    "market_code": "",
                    "market_name": market_name,
                    "sector_17_code": _text(source.get("17業種コード")),
                    "sector_17_name": _text(source.get("17業種区分")),
                    "sector_33_code": _text(source.get("33業種コード")),
                    "sector_33_name": _text(source.get("33業種区分")),
                    "scale_category": _text(source.get("規模区分")),
                    "margin_category": "",
                    "margin_category_name": "",
                    "product_category": "domestic_stock",
                    "provider": "JPX",
                    "retrieved_at_jst": retrieved_at,
                }
            )
        return result
    finally:
        workbook.close()


def _normalize_master(
    rows: list[dict[str, Any]], retrieved_at: str, provider: str
) -> list[dict[str, str]]:
    return [
        {
            "effective_date": _date_text(row.get("Date")),
            "code": _normalize_code(row.get("Code")),
            "company_name": _text(row.get("CoName")),
            "company_name_en": _text(row.get("CoNameEn")),
            "market_code": _text(row.get("Mkt")),
            "market_name": _text(row.get("MktNm")),
            "sector_17_code": _text(row.get("S17")),
            "sector_17_name": _text(row.get("S17Nm")),
            "sector_33_code": _text(row.get("S33")),
            "sector_33_name": _text(row.get("S33Nm")),
            "scale_category": _text(row.get("ScaleCat")),
            "margin_category": _text(row.get("Mrgn")),
            "margin_category_name": _text(row.get("MrgnNm")),
            "product_category": _text(row.get("ProdCat")),
            "provider": provider,
            "retrieved_at_jst": retrieved_at,
        }
        for row in rows
        if _normalize_code(row.get("Code"))
    ]


def _normalize_bars(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    mapping = {
        "date": "Date",
        "code": "Code",
        "open": "O",
        "high": "H",
        "low": "L",
        "close": "C",
        "volume": "Vo",
        "trading_value": "Va",
        "adjustment_factor": "AdjFactor",
        "adjusted_close": "AdjC",
        "adjusted_volume": "AdjVo",
    }
    normalized: list[dict[str, str]] = []
    for row in rows:
        item = {field: _text(row.get(source)) for field, source in mapping.items()}
        item["date"] = _date_text(row.get("Date"))
        item["code"] = _normalize_code(row.get("Code"))
        item["provider"] = "J-Quants API V2"
        if item["date"] and item["code"]:
            normalized.append(item)
    return sorted(normalized, key=lambda item: (item["code"], item["date"]))


def _source_row(
    *, source_id: str, category: str, code: str, title: str,
    published_at: str, retrieved_at: str, url: str, notes: str,
) -> dict[str, str]:
    return {
        "source_id": source_id,
        "category": category,
        "code": code,
        "title": title,
        "published_at_jst": published_at,
        "retrieved_at_jst": retrieved_at,
        "url": url,
        "primary_source": "true",
        "used_for_decision": "true",
        "notes": notes,
    }


def _query_published(scan_date: date, cutoff: datetime) -> str:
    return min(
        cutoff, datetime.combine(scan_date, time(23, 59, 59), tzinfo=JST)
    ).isoformat(timespec="seconds")


def _task(task_id: str, task_type: str, priority: str, reason: str, code: str = "", source_ids: list[str] | None = None) -> dict[str, Any]:
    return {
        "task_id": task_id,
        "task_type": task_type,
        "priority": priority,
        "code": code,
        "reason": reason,
        "source_ids": source_ids or [],
        "status": "PENDING",
        "evidence_source_ids": [],
        "decision_log_path": None,
        "notes": "",
    }


def _manifest_entry(path: Path, run_dir: Path, rows: int, source_url: str) -> dict[str, Any]:
    return {
        "path": path.relative_to(run_dir).as_posix(),
        "sha256": _sha256(path),
        "row_count": rows,
        "source_url": source_url,
        "storage": "PRIVATE_ONLY",
    }


def collect_non_edinet_sources(
    *,
    run_id: str,
    scan_dates: list[date],
    targets: set[str],
    cutoff_at: datetime,
    started: datetime,
    config: dict[str, Any],
    run_dir: Path,
    private: Path,
    fixture_dir: Path | None = None,
    environ: Mapping[str, str] | None = None,
) -> dict[str, Any]:
    """Collect available official reference data and report fail-closed coverage."""

    environment = os.environ if environ is None else environ
    retrieved = started.isoformat(timespec="seconds")
    timeout = int(config.get("request_timeout_seconds", 30))
    raw_dir = run_dir / "raw-sources"
    normalized_dir = run_dir / "reference-data"
    raw_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
    normalized_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
    providers: dict[str, dict[str, Any]] = {}
    sources: list[dict[str, str]] = []
    tasks: list[dict[str, Any]] = []
    blocking_failures: list[dict[str, str]] = []
    successful_gap_sources: set[str] = set()
    manifest: dict[str, Any] = {
        "schema_version": "1.0",
        "run_id": run_id,
        "cutoff_at_jst": cutoff_at.isoformat(timespec="seconds"),
        "retrieved_at_jst": retrieved,
        "datasets": {},
        "notes": [
            "All data files are private-only; public redistribution is prohibited.",
            "Calculated fields are identified separately from official source fields.",
        ],
    }

    def provider(name: str, status: str, requests: int, records: int, error: str = "") -> None:
        providers[name] = {
            "status": status,
            "request_count": requests,
            "record_count": records,
            "error": error,
        }

    # JPX's public monthly list is the credential-free baseline.  A copy is kept
    # for every run so that later research does not accidentally use today's
    # constituents for an older decision.
    jpx_config = config.get("jpx_listed_master", {})
    if not jpx_config.get("enabled", True):
        provider("jpx_listed_master", "DISABLED", 0, 0, "provider disabled")
    else:
        try:
            if fixture_dir:
                rows, pages = _fixture_pages(fixture_dir, f"jpx-listed-{run_id}")
                jpx_rows = _normalize_master(rows, retrieved, "JPX")
                _atomic_json(raw_dir / f"jpx-listed-{run_id}.json", {"pages": pages})
                raw_path = raw_dir / f"jpx-listed-{run_id}.json"
            else:
                raw_path = raw_dir / f"jpx-listed-{run_id}.xlsx"
                _download_jpx_list(str(jpx_config["url"]), raw_path, timeout)
                jpx_rows = _parse_jpx_list(raw_path, retrieved)
                minimum = int(jpx_config.get("minimum_issue_count", 3000))
                if len(jpx_rows) < minimum:
                    raise NonEdinetSourceError(
                        f"JPX listed master has only {len(jpx_rows)} domestic stocks"
                    )
            snapshot = normalized_dir / "jpx-listed-master.csv.gz"
            _atomic_csv_gzip(snapshot, MASTER_FIELDS, jpx_rows)
            manifest["datasets"]["jpx_listed_master"] = {
                **_manifest_entry(snapshot, run_dir, len(jpx_rows), str(jpx_config["url"])),
                "raw_path": raw_path.relative_to(run_dir).as_posix(),
                "raw_sha256": _sha256(raw_path),
                "point_in_time": True,
            }
            provider("jpx_listed_master", "OK", 1, len(jpx_rows))
            sources.append(
                _source_row(
                    source_id=f"jpx-listed-master-{run_id}-{_sha256(snapshot)[:12]}",
                    category="jpx_reference_data",
                    code="",
                    title=f"JPX listed domestic-stock master snapshot ({len(jpx_rows)} issues)",
                    published_at=_query_published(cutoff_at.date(), cutoff_at),
                    retrieved_at=retrieved,
                    url=str(jpx_config["url"]),
                    notes="official point-in-time snapshot; raw and normalized checksums in reference-data/manifest.json",
                )
            )
            successful_gap_sources.add("jpx_security_master")
        except (KeyError, NonEdinetSourceError, ValueError, OSError) as error:
            provider("jpx_listed_master", "ERROR", 1, 0, str(error))
            blocking_failures.append(
                {
                    "source": "jpx_security_master",
                    "impact": "the listed universe is not point-in-time verified; candidate selection is blocked",
                    "reason": "Retry the official JPX listed-master snapshot and attach its checksum evidence",
                }
            )

    jq_config = config.get("jquants", {})
    jq_enabled = bool(jq_config.get("enabled"))
    jq_key = environment.get(str(jq_config.get("api_key_env", "")), "")
    jq_available = jq_enabled and (bool(fixture_dir) or bool(jq_key))
    base_url = str(jq_config.get("base_url", "https://api.jquants.com/v2"))

    if not jq_enabled:
        for name in (
            "jquants_security_master", "jquants_market_data",
            "jquants_financials", "jquants_earnings_calendar", "jquants_trading_calendar",
        ):
            provider(name, "DISABLED", 0, 0, "provider disabled")
    elif not jq_available:
        for name in (
            "jquants_security_master", "jquants_market_data",
            "jquants_financials", "jquants_earnings_calendar", "jquants_trading_calendar",
        ):
            provider(name, "MISSING_CREDENTIAL", 0, 0, "JQUANTS_API_KEY is not set")
    else:
        # Daily point-in-time master, including industry and market classifications.
        if jq_config.get("listed_master_enabled", True):
            try:
                stem = f"jquants-master-{run_id}"
                if fixture_dir:
                    rows, pages = _fixture_pages(fixture_dir, stem)
                else:
                    rows, pages = _jquants_pages(
                        base_url=base_url, path="equities/master",
                        params={"date": run_id}, api_key=jq_key, timeout=timeout,
                    )
                raw_path = raw_dir / f"{stem}.json"
                _atomic_json(raw_path, {"pages": pages})
                normalized = _normalize_master(rows, retrieved, "J-Quants API V2")
                path = normalized_dir / "jquants-security-master.csv.gz"
                _atomic_csv_gzip(path, MASTER_FIELDS, normalized)
                endpoint = f"{base_url}/equities/master"
                manifest["datasets"]["jquants_security_master"] = {
                    **_manifest_entry(path, run_dir, len(normalized), endpoint),
                    "raw_path": raw_path.relative_to(run_dir).as_posix(),
                    "raw_sha256": _sha256(raw_path),
                    "point_in_time": True,
                }
                provider("jquants_security_master", "OK", len(pages), len(rows))
                sources.append(
                    _source_row(
                        source_id=f"jquants-master-{run_id}-{_sha256(path)[:12]}",
                        category="jpx_reference_data", code="",
                        title=f"J-Quants point-in-time listed master ({len(normalized)} issues)",
                        published_at=_query_published(cutoff_at.date(), cutoff_at),
                        retrieved_at=retrieved, url=endpoint,
                        notes="official JPXI API V2; industry and market classifications preserved privately",
                    )
                )
            except (KeyError, NonEdinetSourceError, ValueError, OSError) as error:
                provider("jquants_security_master", "ERROR", 0, 0, str(error))
        else:
            provider("jquants_security_master", "DISABLED", 0, 0, "dataset disabled")

        # Official bars are fetched per active target over enough calendar days
        # to calculate a 20-session liquidity measure.
        all_bars: list[dict[str, Any]] = []
        bar_requests = 0
        bar_error = ""
        if not targets:
            provider("jquants_market_data", "NOT_APPLICABLE", 0, 0)
        elif jq_config.get("daily_bars_enabled", True):
            try:
                lookback = max(30, int(jq_config.get("liquidity_lookback_calendar_days", 45)))
                start_date = (cutoff_at.date() - timedelta(days=lookback)).isoformat()
                for code in sorted(targets):
                    stem = f"jquants-bars-{code}-{run_id}"
                    if fixture_dir:
                        rows, pages = _fixture_pages(fixture_dir, stem)
                    else:
                        rows, pages = _jquants_pages(
                            base_url=base_url, path="equities/bars/daily",
                            params={"code": code, "from": start_date, "to": run_id},
                            api_key=jq_key, timeout=timeout,
                        )
                    bar_requests += len(pages)
                    all_bars.extend(rows)
                    _atomic_json(raw_dir / f"{stem}.json", {"pages": pages})
                normalized_bars = _normalize_bars(all_bars)
                bars_path = normalized_dir / "jquants-daily-bars.csv.gz"
                _atomic_csv_gzip(bars_path, BAR_FIELDS, normalized_bars)
                endpoint = f"{base_url}/equities/bars/daily"
                manifest["datasets"]["jquants_daily_bars"] = _manifest_entry(
                    bars_path, run_dir, len(normalized_bars), endpoint
                )

                liquidity_rows: list[dict[str, str]] = []
                missing: list[str] = []
                minimum_sessions = max(1, int(jq_config.get("minimum_liquidity_sessions", 20)))
                maximum_age = max(0, int(jq_config.get("maximum_official_price_age_days", 7)))
                for code in sorted(targets):
                    code_rows = [row for row in normalized_bars if row["code"] == code]
                    eligible = [
                        row for row in code_rows
                        if _number(row["trading_value"]) is not None
                        and float(row["trading_value"]) > 0
                        and date.fromisoformat(row["date"]) <= cutoff_at.date()
                    ]
                    eligible = sorted(eligible, key=lambda row: row["date"])[-minimum_sessions:]
                    if len(eligible) < minimum_sessions:
                        missing.append(f"{code}: {len(eligible)}/{minimum_sessions} sessions")
                        continue
                    latest = eligible[-1]
                    if (cutoff_at.date() - date.fromisoformat(latest["date"])).days > maximum_age:
                        missing.append(f"{code}: latest official bar is stale ({latest['date']})")
                        continue
                    values = [float(row["trading_value"]) for row in eligible]
                    liquidity_rows.append(
                        {
                            "as_of_date": latest["date"], "code": code,
                            "sessions": str(len(eligible)),
                            "average_trading_value_20d": format(statistics.fmean(values), ".12g"),
                            "median_trading_value_20d": format(statistics.median(values), ".12g"),
                            "latest_close": latest["close"], "latest_volume": latest["volume"],
                            "latest_trading_value": latest["trading_value"],
                            "provider": "J-Quants API V2",
                            "evidence_path": bars_path.relative_to(run_dir).as_posix(),
                        }
                    )
                    sources.append(
                        _source_row(
                            source_id=f"jquants-price-{latest['date']}-{code}",
                            category="market_data", code=code,
                            title=f"Official daily price and 20-session liquidity for {code}",
                            published_at=f"{latest['date']}T15:30:00+09:00",
                            retrieved_at=retrieved, url=endpoint,
                            notes="official source fields plus calculated 20-session average and median trading value",
                        )
                    )
                liquidity_path = normalized_dir / "liquidity-20d.csv"
                _atomic_csv(liquidity_path, LIQUIDITY_FIELDS, liquidity_rows)
                manifest["datasets"]["liquidity_20d"] = {
                    **_manifest_entry(liquidity_path, run_dir, len(liquidity_rows), endpoint),
                    "calculation": "last 20 positive official daily trading-value observations at or before cutoff",
                }
                if missing:
                    bar_error = "; ".join(missing)
                    provider("jquants_market_data", "INCOMPLETE", bar_requests, len(all_bars), bar_error)
                else:
                    provider("jquants_market_data", "OK", bar_requests, len(all_bars))
                    successful_gap_sources.add("official_market_data")

                action_rows: list[dict[str, str]] = []
                for row in normalized_bars:
                    factor = _number(row["adjustment_factor"])
                    if factor is None or math.isclose(factor, 1.0):
                        continue
                    action_id = f"jq-adjustment-{row['code']}-{row['date']}-{_stable_id(row['adjustment_factor'])}"
                    action_source_id = f"jquants-adjustment-{row['date']}-{row['code']}"
                    sources.append(
                        _source_row(
                            source_id=action_source_id,
                            category="market_data",
                            code=row["code"],
                            title=f"Official adjustment factor {row['adjustment_factor']} for {row['code']}",
                            published_at=f"{row['date']}T15:30:00+09:00",
                            retrieved_at=retrieved,
                            url=endpoint,
                            notes="factor is official; corporate-action type and ratio remain subject to company-IR confirmation",
                        )
                    )
                    action_rows.append(
                        {
                            "action_id": action_id, "announced_at_jst": "",
                            "effective_date": row["date"], "code": row["code"],
                            "action_type": "JQUANTS_ADJUSTMENT_FACTOR",
                            "ratio_numerator": "", "ratio_denominator": "",
                            "old_code": "", "new_code": "",
                            "normalization_factor": row["adjustment_factor"],
                            "source_url": endpoint, "applied_at_jst": "", "applied_by": "",
                            "evidence_path": bars_path.relative_to(run_dir).as_posix(),
                            "notes": "Official adjustment factor detected; action type and ratio require company-IR confirmation before application.",
                        }
                    )
                    tasks.append(
                        _task(
                            task_id=f"review-{action_id}", task_type="CORPORATE_ACTION_REVIEW",
                            priority="URGENT", code=row["code"],
                            reason="Official adjustment factor changed; confirm split/consolidation details with company IR",
                            source_ids=[action_source_id],
                        )
                    )
                _append_csv_unique(
                    private / "corporate-actions.csv",
                    tuple(action_rows[0].keys()) if action_rows else (
                        "action_id", "announced_at_jst", "effective_date", "code", "action_type",
                        "ratio_numerator", "ratio_denominator", "old_code", "new_code",
                        "normalization_factor", "source_url", "applied_at_jst", "applied_by",
                        "evidence_path", "notes",
                    ),
                    action_rows,
                    "action_id",
                )
            except (KeyError, NonEdinetSourceError, ValueError, OSError) as error:
                bar_error = str(error)
                provider("jquants_market_data", "ERROR", bar_requests, len(all_bars), bar_error)
        else:
            provider("jquants_market_data", "DISABLED", 0, 0, "dataset disabled")

        # Financial summaries preserve both forecast revisions and the reported
        # fiscal-year share count used as a starting point for dilution review.
        financial_rows: list[dict[str, Any]] = []
        financial_requests = 0
        if jq_config.get("financial_summary_enabled", True):
            try:
                for scan_date in scan_dates:
                    stem = f"jquants-financials-{scan_date.isoformat()}"
                    if fixture_dir:
                        rows, pages = _fixture_pages(fixture_dir, stem)
                    else:
                        rows, pages = _jquants_pages(
                            base_url=base_url, path="fins/summary",
                            params={"date": scan_date.isoformat()}, api_key=jq_key, timeout=timeout,
                        )
                    financial_requests += len(pages)
                    financial_rows.extend(rows)
                    _atomic_json(raw_dir / f"{stem}.json", {"pages": pages})
                endpoint = f"{base_url}/fins/summary"
                accepted = []
                for row in financial_rows:
                    disclosed = _disclosed_at(row)
                    if datetime.fromisoformat(disclosed) <= cutoff_at:
                        accepted.append((row, disclosed))
                forecast_rows: list[dict[str, str]] = []
                share_rows: list[dict[str, str]] = []
                for row, disclosed in accepted:
                    code = _normalize_code(row.get("Code"))
                    disclosure_number = _text(row.get("DiscNo"))
                    source_id = f"jq-fin-{disclosure_number or _stable_id(disclosed, code, json.dumps(row, sort_keys=True))}"
                    forecast_rows.append(
                        {
                            "source_id": source_id, "disclosed_at_jst": disclosed, "code": code,
                            "disclosure_number": disclosure_number, "document_type": _text(row.get("DocType")),
                            "current_period_type": _text(row.get("CurPerType")),
                            "current_fiscal_year_end": _date_text(row.get("CurFYEn")),
                            "forecast_sales": _text(row.get("FSales")),
                            "forecast_operating_profit": _text(row.get("FOP")),
                            "forecast_ordinary_profit": _text(row.get("FOdP")),
                            "forecast_net_profit": _text(row.get("FNP")),
                            "forecast_eps": _text(row.get("FEPS")),
                            "next_forecast_sales": _text(row.get("NxFSales")),
                            "next_forecast_operating_profit": _text(row.get("NxFOP")),
                            "next_forecast_ordinary_profit": _text(row.get("NxFOdP")),
                            "next_forecast_net_profit": _text(row.get("NxFNp")),
                            "next_forecast_eps": _text(row.get("NxFEPS")),
                            "material_change_submitted": _text(row.get("MatChgSub")),
                            "retrieved_at_jst": retrieved,
                            "evidence_path": f"raw-sources/jquants-financials-{_date_text(row.get('DiscDate'))}.json",
                        }
                    )
                    if any(_text(row.get(field)) for field in ("ShOutFY", "TrShFY", "AvgSh")):
                        share_rows.append(
                            {
                                "source_id": source_id, "disclosed_at_jst": disclosed,
                                "fiscal_year_end": _date_text(row.get("CurFYEn")), "code": code,
                                "issued_shares_fy": _text(row.get("ShOutFY")),
                                "treasury_shares_fy": _text(row.get("TrShFY")),
                                "average_shares": _text(row.get("AvgSh")),
                                "retrieved_at_jst": retrieved,
                                "evidence_path": f"raw-sources/jquants-financials-{_date_text(row.get('DiscDate'))}.json",
                                "notes": "Fiscal-year disclosure values; not a real-time issued-share count.",
                            }
                        )
                    if code in targets:
                        sources.append(
                            _source_row(
                                source_id=source_id, category="jquants_financials", code=code,
                                title=f"J-Quants financial summary {_text(row.get('DocType'))}",
                                published_at=disclosed, retrieved_at=retrieved, url=endpoint,
                                notes="official financial summary; forecast and reported share-count history saved privately",
                            )
                        )
                        tasks.append(
                            _task(
                                task_id=f"review-{source_id}", task_type="FINANCIAL_REVISION_REVIEW",
                                priority="HIGH", code=code,
                                reason="New official financial summary for an active target",
                                source_ids=[source_id],
                            )
                        )
                forecast_path = normalized_dir / "forecast-revisions.csv"
                share_path = normalized_dir / "share-counts.csv"
                _atomic_csv(forecast_path, FORECAST_FIELDS, forecast_rows)
                _atomic_csv(share_path, SHARE_FIELDS, share_rows)
                manifest["datasets"]["forecast_revisions"] = _manifest_entry(
                    forecast_path, run_dir, len(forecast_rows), endpoint
                )
                manifest["datasets"]["share_counts"] = {
                    **_manifest_entry(share_path, run_dir, len(share_rows), endpoint),
                    "limitation": "reported fiscal-year values; current dilution securities still require company-IR review",
                }
                _append_csv_unique(private / "forecast-history.csv", FORECAST_FIELDS, forecast_rows, "source_id")
                _append_csv_unique(private / "share-count-history.csv", SHARE_FIELDS, share_rows, "source_id")
                provider("jquants_financials", "OK", financial_requests, len(financial_rows))
            except (KeyError, NonEdinetSourceError, ValueError, OSError) as error:
                provider("jquants_financials", "ERROR", financial_requests, len(financial_rows), str(error))
        else:
            provider("jquants_financials", "DISABLED", 0, 0, "dataset disabled")

        earnings_rows: list[dict[str, Any]] = []
        earnings_requests = 0
        if jq_config.get("earnings_calendar_enabled", True):
            try:
                for scan_date in scan_dates:
                    stem = f"jquants-earnings-{scan_date.isoformat()}"
                    if fixture_dir:
                        rows, pages = _fixture_pages(fixture_dir, stem)
                    else:
                        rows, pages = _jquants_pages(
                            base_url=base_url, path="fins/earnings-date",
                            params={"date": scan_date.isoformat()}, api_key=jq_key, timeout=timeout,
                        )
                    earnings_requests += len(pages)
                    earnings_rows.extend(rows)
                    _atomic_json(raw_dir / f"{stem}.json", {"pages": pages})
                endpoint = f"{base_url}/fins/earnings-date"
                normalized_earnings: list[dict[str, str]] = []
                for row in earnings_rows:
                    code = _normalize_code(row.get("Code"))
                    published = _date_text(row.get("PubDate"))
                    source_id = f"jq-earnings-{_stable_id(published, code, _text(row.get('FQName')), _date_text(row.get('SchDate')))}"
                    normalized_earnings.append(
                        {
                            "source_id": source_id, "published_date": published,
                            "scheduled_date": _date_text(row.get("SchDate")),
                            "fiscal_quarter": _text(row.get("FQName")),
                            "fiscal_year_end_mmdd": _text(row.get("FYE")), "code": code,
                            "company_name": _text(row.get("CoName")),
                            "company_name_en": _text(row.get("CoNameEn")),
                            "retrieved_at_jst": retrieved,
                            "evidence_path": f"raw-sources/jquants-earnings-{published}.json",
                        }
                    )
                    if code in targets:
                        sources.append(
                            _source_row(
                                source_id=source_id, category="earnings_calendar", code=code,
                                title=f"Earnings schedule {_text(row.get('FQName'))}: {_date_text(row.get('SchDate')) or 'undecided'}",
                                published_at=f"{published}T00:00:00+09:00", retrieved_at=retrieved,
                                url=endpoint, notes="official schedule history; prior records are retained after a change",
                            )
                        )
                        tasks.append(
                            _task(
                                task_id=f"review-{source_id}", task_type="EARNINGS_SCHEDULE_REVIEW",
                                priority="HIGH", code=code,
                                reason="Earnings announcement schedule was published or changed",
                                source_ids=[source_id],
                            )
                        )
                path = normalized_dir / "earnings-calendar-changes.csv"
                _atomic_csv(path, EARNINGS_FIELDS, normalized_earnings)
                manifest["datasets"]["earnings_calendar_changes"] = _manifest_entry(
                    path, run_dir, len(normalized_earnings), endpoint
                )
                _append_csv_unique(private / "earnings-calendar-history.csv", EARNINGS_FIELDS, normalized_earnings, "source_id")
                provider("jquants_earnings_calendar", "OK", earnings_requests, len(earnings_rows))
            except (KeyError, NonEdinetSourceError, ValueError, OSError) as error:
                provider("jquants_earnings_calendar", "ERROR", earnings_requests, len(earnings_rows), str(error))
        else:
            provider("jquants_earnings_calendar", "DISABLED", 0, 0, "dataset disabled")

        if jq_config.get("trading_calendar_enabled", True):
            try:
                forward = max(7, int(jq_config.get("calendar_forward_days", 21)))
                stem = f"jquants-calendar-{run_id}"
                if fixture_dir:
                    rows, pages = _fixture_pages(fixture_dir, stem)
                else:
                    rows, pages = _jquants_pages(
                        base_url=base_url, path="markets/calendar",
                        params={"from": run_id, "to": (cutoff_at.date() + timedelta(days=forward)).isoformat()},
                        api_key=jq_key, timeout=timeout,
                    )
                raw_path = raw_dir / f"{stem}.json"
                _atomic_json(raw_path, {"pages": pages})
                calendar_rows = [
                    {"date": _date_text(row.get("Date")), "holiday_division": _text(row.get("HolDiv"))}
                    for row in rows if _date_text(row.get("Date"))
                ]
                _atomic_json(
                    run_dir / "trading-calendar.json",
                    {
                        "schema_version": "1.0", "provider": "J-Quants API V2",
                        "retrieved_at_jst": retrieved, "rows": calendar_rows,
                    },
                )
                endpoint = f"{base_url}/markets/calendar"
                next_sessions = [
                    row for row in calendar_rows
                    if date.fromisoformat(row["date"]) > cutoff_at.date()
                    and row["holiday_division"] in {"1", "2"}
                ]
                if not next_sessions:
                    raise NonEdinetSourceError("trading calendar has no future cash-equity session")
                manifest["datasets"]["trading_calendar"] = {
                    **_manifest_entry(run_dir / "trading-calendar.json", run_dir, len(calendar_rows), endpoint),
                    "raw_path": raw_path.relative_to(run_dir).as_posix(),
                    "raw_sha256": _sha256(raw_path),
                }
                provider("jquants_trading_calendar", "OK", len(pages), len(rows))
                successful_gap_sources.add("trading_calendar")
                sources.append(
                    _source_row(
                        source_id=f"jquants-calendar-{run_id}", category="trading_calendar", code="",
                        title=f"Official trading calendar through {calendar_rows[-1]['date']}",
                        published_at=_query_published(cutoff_at.date(), cutoff_at),
                        retrieved_at=retrieved, url=endpoint,
                        notes="official JPX/OSE holiday divisions; cash-equity next session derived from HolDiv 1/2",
                    )
                )
            except (KeyError, NonEdinetSourceError, ValueError, OSError) as error:
                provider("jquants_trading_calendar", "ERROR", 0, 0, str(error))
        else:
            provider("jquants_trading_calendar", "DISABLED", 0, 0, "dataset disabled")

    _atomic_json(normalized_dir / "manifest.json", manifest)
    return {
        "providers": providers,
        "source_rows": sources,
        "tasks": tasks,
        "blocking_failures": blocking_failures,
        "successful_gap_sources": successful_gap_sources,
        "manifest_path": "reference-data/manifest.json",
    }
